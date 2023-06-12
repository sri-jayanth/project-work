import struct
from openpyxl import Workbook

def read_dat_file(file_path):
    with open(file_path, "rb") as file:
        return file.read()

def extract_header(data):
    header_struct = struct.Struct(">22s")
    header_data = data[:header_struct.size]
    return header_struct.unpack(header_data)[0].decode("utf-8")

def extract_packet2(data):
    packet2_struct = struct.Struct(">B4x16d3x16d3x6d11x23d")
    packet2_start = 47
    packet2_data = data[packet2_start:packet2_start + packet2_struct.size]
    packet2_values = list(packet2_struct.unpack(packet2_data))
    return packet2_values

def convert_packet2_data(packet2_values):
    packet2_values[0] = str(packet2_values[0])  # Convert packet validity byte to string
    byte_skips = [46, 180, 183, 312, 315, 388, 391, 449, 699, 709]
    for skip in byte_skips:
        packet2_values = packet2_values[:skip] + packet2_values[skip + 1:]
    return packet2_values

def write_to_excel(data, header, packet2_values, sheet):
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=row - 2)  # SI.NO
    sheet.cell(row=row, column=2, value=header)  # Frame ID
    for col, value in enumerate(packet2_values, start=3):
        sheet.cell(row=row, column=col, value=value)

def process_dat_file(file_path):
    data = read_dat_file(file_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="PACKET 2 DATA")
    headings = ["SI.NO", "Frame ID", "Packet validity byte", "Parameter 1", "Parameter 2", "Parameter 3"]
    for col, heading in enumerate(headings, start=1):
        sheet.cell(row=2, column=col, value=heading)
    frame_size = 1998
    frame_count = len(data) // frame_size
    for frame_id in range(frame_count):
        start_index = frame_id * frame_size
        header = extract_header(data[start_index:])
        packet2_values = extract_packet2(data[start_index:])
        packet2_values = convert_packet2_data(packet2_values)
        write_to_excel(data, header, packet2_values, sheet)
    workbook.save("RESULTS.xlsx")

# Usage example
process_dat_file("data.dat")
