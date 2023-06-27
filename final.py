import openpyxl
import tkinter
import struct
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import Tk, Label, StringVar, OptionMenu,Button

def data_processing():
    # Function to convert bytes to double precision floating point number (big endian)
    def bytes_to_double(bytes_data):
        return int.from_bytes(bytes_data, 'big') / (2 ** 64)

    # Function to convert bytes to binary number (1 byte)
    def bytes_to_binary(bytes_data):
        return bin(int.from_bytes(bytes_data, 'big'))[2:].zfill(8)

    # Function to convert bytes to unsigned integer (2 bytes)
    def bytes_to_unsigned_int(bytes_data):
        return int.from_bytes(bytes_data)

    # Function to convert bytes to string (header)
    def bytes_to_string(bytes_data):
        return bytes_data.decode('utf-8')

    datafile = 'C:\data\project' # replace with your file directory
    # Read the .dat file as binary
    with open(datafile, 'rb') as file:
        data = file.read()

    # Create a new Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the heading
    sheet.cell(row=1, column=1, value='PACKET 2 DATA')

    # Write column headings
    headings = ['Frame Number', 'Heading', 'Packet Validity Byte']
    parameters = ['Parameter ' + str(i) for i in range(1, 104)]
    column = 3

    for heading in headings + parameters:                         
        sheet.cell(row=3, column=column, value=heading)
        column += 1

    # Extract data from each frame
    frame_size = 1998
    header_size = 22
    packet1_size = 24
    packet2_size = 842
    reserved_data_size = 101
    total_parameters = 103

    frame_count = len(data) // frame_size

    for frame in range(frame_count):
        start = frame * frame_size

        # Extract header
        header_data = data[start:start + header_size]
        header = bytes_to_string(header_data)

        # Extract packet 2
        packet2_start = start + header_size + packet1_size
        packet2_data = data[packet2_start:packet2_start + packet2_size]
        packet_validity = bytes_to_binary(packet2_data[:1])
        parameters_data = packet2_data[1:]

        # Extract parameters
        parameters = []
        index = 0

        for _ in range(total_parameters):                                          
            for i in range(16):
                parameter = int.from_bytes(frame[index:index + 8], 'big')
                parameters.append(parameter)
                index += 8

            # Skip 3 bytes
            index += 3

            for _ in range(9):
                parameter = int.from_bytes(frame[index:index + 8], 'big')
                parameters.append(parameter)
                index += 8

            # Skip 9 bytes
            index += 9

            # Parameters with 8 bytes each
            for _ in range(31):
                parameter = int.from_bytes(frame[index:index + 8], 'big')
                parameters.append(parameter)
                index += 8

            # Skip 11 bytes
            index += 11

            # Parameters with 8 bytes each
            for _ in range(23):
                parameter = int.from_bytes(frame[index:index + 8], 'big')
                parameters.append(parameter)
                index += 8

            parameters.append(total_parameters)
            index += 8

        # Write data to Excel
        row = frame + 4  # Adjust for header and column headings
        sheet.cell(row=row, column=1, value=frame + 1)
        sheet.cell(row=row, column=2, value=header)
        sheet.cell(row=row, column=3, value=packet_validity)

        column = 4
        for parameter in parameters:
            sheet.cell(row=row, column=column, value=parameter)
            column += 1

    # Save the Excel file
    workbook.save('RESULTS.xlsx')


def gui():
    import tkinter as tk
    import matplotlib.pyplot as plt
    import pandas as pd

    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('your_file.xlsx')

    # Create the GUI window
    window = tk.Tk()
    window.title("Time Series Graph")
    window.geometry("400x200")

    # Get the list of parameters from the first row of the DataFrame
    parameters = list(df.columns)

    # Create the dropdown menu
    selected_parameter = tk.StringVar(window)
    selected_parameter.set(parameters[0])
    parameter_dropdown = tk.OptionMenu(window, selected_parameter, *parameters)
    parameter_dropdown.pack()

    # Function to plot the graph
    def plot_graph():
        parameter = selected_parameter.get()
        data = df[parameter].values
        time = range(len(data))

        plt.plot(time, data)
        plt.xlabel("Time")
        plt.ylabel(parameter)
        plt.title("Time Series Graph")
        plt.show()

    # Create the plot button
    plot_button = tk.Button(window, text="Plot Graph", command=plot_graph)
    plot_button.pack()

    # Start the GUI event loop
    window.mainloop()

if __name__ == "__main__":
    data_processing()
    gui()
