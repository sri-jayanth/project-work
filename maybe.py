import openpyxl

# Function to convert bytes to double precision floating point number (big endian)
def bytes_to_double(bytes_data):
    return int.from_bytes(bytes_data, 'big') / (2 ** 64)

# Function to convert bytes to binary number (1 byte)
def bytes_to_binary(bytes_data):
    return bin(int.from_bytes(bytes_data, 'big'))[2:].zfill(8)

# Function to convert bytes to unsigned integer (2 bytes)
def bytes_to_unsigned_int(bytes_data):
    return int.from_bytes(bytes_data, 'big')

# Function to convert bytes to string (header)
def bytes_to_string(bytes_data):
    return bytes_data.decode('utf-8')

# Read the .dat file as binary
with open('data.dat', 'rb') as file:
    data = file.read()

# Create a new Excel file
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the heading
sheet.cell(row=1, column=1, value='PACKET 2 DATA')

# Write column headings
headings = ['Frame Number', 'Heading', 'Packet Validity Byte']
parameters = ['Parameter ' + str(i) for i in range(1, 104)]
column = 3

for heading in headings + parameters:
    sheet.cell(row=3, column=column, value=heading)
    column += 1

# Extract data from each frame
frame_size = 1998
header_size = 22
packet1_size = 24
packet2_size = 842
reserved_data_size = 101
total_parameters = 103

frame_count = len(data) // frame_size

for frame in range(frame_count):
    start = frame * frame_size

    # Extract header
    header_data = data[start:start + header_size]
    header = bytes_to_string(header_data)

    # Extract packet 2
    packet2_start = start + header_size + packet1_size
    packet2_data = data[packet2_start:packet2_start + packet2_size]
    packet_validity = bytes_to_binary(packet2_data[:1])
    parameters_data = packet2_data[1:]

    # Extract parameters
    parameters = []
    index = 0

    for _ in range(total_parameters):
        if index in range(52, 181) or index in range(184, 313) or index in range(316, 389) or index in range(392, 441) or index in range(450, 710) or index in range(894, 899):
            parameter_data = parameters_data[index:index + 8]
            parameter_value = bytes_to_double(parameter_data)
        elif index in range(181, 184) or index in range(313, 316) or index in range(441, 450) or index in range(710, 894):
            index += 3
            continue
        elif index in range(899, 910):
            index += 11
            continue
        else:
            parameter_data = parameters_data[index:index + 8]
            parameter_value = bytes_to_unsigned_int(parameter_data)

        parameters.append(parameter_value)
        index += 8

    # Write data to Excel
    row = frame + 4  # Adjust for header and column headings
    sheet.cell(row=row, column=1, value=frame + 1)
    sheet.cell(row=row, column=2, value=header)
    sheet.cell(row=row, column=3, value=packet_validity)

    column = 4
    for parameter in parameters:
        sheet.cell(row=row, column=column, value=parameter)
        column += 1

# Save the Excel file
workbook.save('RESULTS.xlsx')
